import openpyxl
from tkinter import *


root = Tk()
text = Text(root, height=2, width=10, font="Arial 14", wrap=WORD)
button = Button(root, height=2, width=8, font="Arial 14", text="Создать")


def new_blank(event=None):
    index = text.get('1.0', END)
    file = openpyxl.load_workbook("Data/Base.xlsx")
    sheet = file["Общий"]

    new_file = openpyxl.load_workbook("Data/Blank.xlsx")
    new_file_sheet = new_file.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == int(index):
            new_file_sheet.cell(row=5, column=16, value=row[2]) # лицевой
            new_file_sheet.cell(row=8, column=33, value=row[1]) # ФИО
            new_file_sheet.cell(row=19, column=1, value="произвел инвентаризацию газифицированного домовладения по адресу: " + row[3]) # Адрес
            new_file_sheet.cell(row=23, column=57, value=row[9]) # количество проживающих
            new_file_sheet.cell(row=29, column=57, value=row[4]) # прибор учета
            new_file_sheet.cell(row=30, column=57, value=row[5]) # номер счетчика
            new_file_sheet.cell(row=31, column=57, value=row[8]) # Дата выпуска
            new_file_sheet.cell(row=32, column=57, value=row[7]) # Дата поверки
            new_file_sheet.cell(row=33, column=57, value=row[6]) # Дата очередной поверки
            new_file.save(f"Сгенерированные бланки/{row[2]}.xlsx")
            break

button.bind('<Button-1>', new_blank)
text.pack()
button.pack()
root.mainloop()
